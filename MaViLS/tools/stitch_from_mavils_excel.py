# E:\Code\MaViLS\tools\stitch_from_mavils_excel.py
# -*- coding: utf-8 -*-
"""
Stitch SRT transcripts into LaTeX pages using MaViLS Key/Value excel mapping.

Inputs:
  --in_tex   merged_1-4.tex (with page markers like: % <base>_page_<n>)
  --xlsx     one or more MaViLS excels (Key=seconds, Value=1-based page index)
  --srt      one or more SRT files (same count as xlsx; each is one video transcript)
  --out_tex  output LaTeX

Key features:
  - Parses Key/Value xlsx (Sheet1) into continuous, gap-free spans.
  - Maps Value (1-based) -> page_label list derived from in_tex markers.
  - Assigns each SRT segment to the span where its midpoint time falls.
  - Writes ALL matched SRT segments into the corresponding page (no truncation, no "...").
"""

import argparse
import math
import os
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import openpyxl

# -------------------- LaTeX page markers --------------------
PAGE_MARKER_RE = re.compile(r"^%\s*(.*?)_page_(\d+)\s*$", flags=re.MULTILINE)

# -------------------- Utils --------------------
SPECIAL_MAP = {
    '\\': r'\textbackslash{}',
    '&':  r'\&',
    '%':  r'\%',
    '$':  r'\$',
    '#':  r'\#',
    '_':  r'\_',
    '{':  r'\{',
    '}':  r'\}',
    '~':  r'\textasciitilde{}',
    '^':  r'\textasciicircum{}',
}

def latex_escape(s: str) -> str:
    return ''.join(SPECIAL_MAP.get(ch, ch) for ch in s)

def fmt_time(sec: float) -> str:
    if sec is None:
        return "??:??"
    if sec < 0:
        sec = 0.0
    s = int(sec)
    ms = int(round((sec - s) * 1000))
    h = s // 3600
    m = (s % 3600) // 60
    ss = s % 60
    if h > 0:
        return f"{h:02d}:{m:02d}:{ss:02d}.{ms:03d}"
    return f"{m:02d}:{ss:02d}.{ms:03d}"

def median(xs: List[float], default: float = 1.0) -> float:
    xs = [x for x in xs if x is not None and x > 1e-9]
    if not xs:
        return default
    xs.sort()
    return xs[len(xs)//2]

# -------------------- Parse SRT --------------------
SRT_TS_RE = re.compile(
    r"(\d{2}):(\d{2}):(\d{2}),(\d{3})\s*-->\s*(\d{2}):(\d{2}):(\d{2}),(\d{3})"
)

def _ts_to_sec(hh, mm, ss, ms) -> float:
    return int(hh) * 3600 + int(mm) * 60 + int(ss) + int(ms) / 1000.0

def parse_srt(path: str) -> List[Dict]:
    text = Path(path).read_text(encoding="utf-8", errors="replace")
    blocks = re.split(r"\n\s*\n", text.strip())
    out = []
    for b in blocks:
        lines = [ln.rstrip("\r") for ln in b.splitlines() if ln.strip() != ""]
        if len(lines) < 2:
            continue

        idx = None
        # first line might be index
        if re.fullmatch(r"\d+", lines[0].strip()):
            idx = int(lines[0].strip())
            time_line = lines[1]
            text_lines = lines[2:]
        else:
            time_line = lines[0]
            text_lines = lines[1:]

        m = SRT_TS_RE.search(time_line)
        if not m:
            continue

        t0 = _ts_to_sec(m.group(1), m.group(2), m.group(3), m.group(4))
        t1 = _ts_to_sec(m.group(5), m.group(6), m.group(7), m.group(8))
        seg_text = " ".join([ln.strip() for ln in text_lines]).strip()

        out.append({
            "idx": idx,
            "t_start": float(t0),
            "t_end": float(t1),
            "text": seg_text,
        })
    out.sort(key=lambda x: (x["t_start"], x["t_end"]))
    return out

# -------------------- Parse Key/Value XLSX --------------------
def read_kv_xlsx(path: str) -> List[Tuple[float, int]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = [str(x).strip() if x is not None else "" for x in header]

    if "Key" not in header or "Value" not in header:
        wb.close()
        raise RuntimeError(f"[XLSX] Expected columns Key/Value, got header={header[:10]} in {path}")

    k_idx = header.index("Key")
    v_idx = header.index("Value")

    rows: List[Tuple[float, int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        k = row[k_idx] if k_idx < len(row) else None
        v = row[v_idx] if v_idx < len(row) else None
        if k is None or v is None:
            continue
        try:
            k = float(k)
            v = int(v)
        except Exception:
            continue
        rows.append((k, v))

    wb.close()
    rows.sort(key=lambda x: x[0])
    return rows

def kv_to_spans(rows: List[Tuple[float, int]], video_end: Optional[float]) -> Tuple[List[Dict], float]:
    """
    Convert (t, val) samples into continuous spans (gap-free):
      span i: [start, next_change_time)
    Starts from t=0 using the first value.
    Ends at max(video_end, last_t + median_step).
    """
    if not rows:
        return [], 1.0

    diffs = [rows[i+1][0] - rows[i][0] for i in range(len(rows)-1)]
    step = median(diffs, default=1.0)

    spans: List[Dict] = []

    cur_val = rows[0][1]
    cur_start = 0.0

    for (t, val) in rows[1:]:
        if val != cur_val:
            # end previous span at first time new val is observed
            spans.append({"value": cur_val, "t_start": cur_start, "t_end": float(t)})
            cur_val = val
            cur_start = float(t)

    last_t = float(rows[-1][0])
    end_t = last_t + step
    if video_end is not None:
        end_t = max(end_t, float(video_end))
    spans.append({"value": cur_val, "t_start": cur_start, "t_end": float(end_t)})

    # ensure monotonic non-decreasing (no overlaps / no negative)
    fixed = []
    prev_end = 0.0
    for sp in spans:
        s = max(float(sp["t_start"]), prev_end)
        e = max(float(sp["t_end"]), s)
        fixed.append({"value": int(sp["value"]), "t_start": s, "t_end": e})
        prev_end = e
    return fixed, step

# -------------------- Parse LaTeX pages --------------------
def read_tex_and_page_labels(tex_path: str) -> Tuple[str, List[str]]:
    tex = Path(tex_path).read_text(encoding="utf-8", errors="replace")
    labels: List[str] = []
    for m in PAGE_MARKER_RE.finditer(tex):
        base = m.group(1)
        num = int(m.group(2))
        labels.append(f"{base}_page_{num}")
    return tex, labels

def lecture_no_from_label(label: str) -> Optional[int]:
    # accepts Lecture 01_... or Lecture_02_...
    m = re.search(r"Lecture[\s_]*0*(\d+)", label)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

def build_page_groups(all_labels: List[str]) -> Dict[int, List[str]]:
    groups: Dict[int, List[str]] = {}
    for lab in all_labels:
        no = lecture_no_from_label(lab)
        if no is None:
            continue
        groups.setdefault(no, []).append(lab)
    return groups

def allocate_pages_for_sources(all_labels: List[str], max_values: List[int]) -> List[List[str]]:
    """
    Prefer:
      src0 -> Lecture 1+2
      src1 -> Lecture 3+4
    If not possible, fallback to sequential slicing by max_values.
    """
    groups = build_page_groups(all_labels)
    out: List[List[str]] = []

    if len(max_values) == 2 and (1 in groups or 2 in groups or 3 in groups or 4 in groups):
        cand0 = groups.get(1, []) + groups.get(2, [])
        cand1 = groups.get(3, []) + groups.get(4, [])
        if len(cand0) >= max_values[0] and len(cand1) >= max_values[1]:
            out.append(cand0)
            out.append(cand1)
            return out
        # if grouping exists but not enough, fall back

    # fallback: take pages in order
    idx = 0
    for mv in max_values:
        chunk = all_labels[idx: idx + mv]
        out.append(chunk)
        idx += mv
    return out

# -------------------- Assign SRT segments to spans -> pages --------------------
def assign_segments_to_pages(
    segments: List[Dict],
    spans: List[Dict],
    value_to_label: Dict[int, str],
) -> Tuple[Dict[str, List[Dict]], List[Dict]]:
    """
    Assign each segment to span where midpoint falls.
    """
    page2segs: Dict[str, List[Dict]] = {}
    offslide: List[Dict] = []

    if not spans:
        for s in segments:
            offslide.append(s)
        return page2segs, offslide

    spans_sorted = sorted(spans, key=lambda x: x["t_start"])
    j = 0
    for seg in segments:
        mid = 0.5 * (seg["t_start"] + seg["t_end"])

        # move span pointer
        while j + 1 < len(spans_sorted) and mid >= spans_sorted[j]["t_end"]:
            j += 1

        sp = spans_sorted[j]
        if sp["t_start"] <= mid < sp["t_end"]:
            val = int(sp["value"])
            label = value_to_label.get(val)
            if label:
                page2segs.setdefault(label, []).append(seg)
            else:
                offslide.append(seg)
        else:
            offslide.append(seg)

    # sort segments per page
    for lab in page2segs:
        page2segs[lab].sort(key=lambda x: (x["t_start"], x["t_end"]))
    return page2segs, offslide

# -------------------- Write back into LaTeX --------------------
def stitch_into_tex(
    tex: str,
    page2segs_allsrc: List[Dict[str, List[Dict]]],
    page2spaninfo_allsrc: List[Dict[str, List[Tuple[float, float]]]],
) -> str:
    """
    Insert ASR blocks after each page marker and its following \\clearpage (if found).
    """
    out_parts: List[str] = []
    last = 0

    for m in PAGE_MARKER_RE.finditer(tex):
        out_parts.append(tex[last:m.end()])

        after_marker_pos = m.end()

        # find the first \clearpage after marker (search locally to reduce surprises)
        clear_m = re.search(r"\\clearpage", tex[after_marker_pos:after_marker_pos+4000])
        insert_pos = after_marker_pos
        if clear_m:
            insert_pos = after_marker_pos + clear_m.end()

        out_parts.append(tex[after_marker_pos:insert_pos])

        base = m.group(1)
        num = int(m.group(2))
        label = f"{base}_page_{num}"

        # gather segments from all sources for this label
        segs_merged: List[Tuple[int, Dict]] = []
        span_ranges: List[str] = []

        for src_i, page2segs in enumerate(page2segs_allsrc):
            segs = page2segs.get(label, [])
            if segs:
                for s in segs:
                    segs_merged.append((src_i, s))
            # also show span time ranges if available
            sr = page2spaninfo_allsrc[src_i].get(label, [])
            for (t0, t1) in sr:
                span_ranges.append(f"src={src_i} {fmt_time(t0)}–{fmt_time(t1)}")

        segs_merged.sort(key=lambda x: (x[1]["t_start"], x[1]["t_end"]))

        if segs_merged:
            # header comment
            hdr = " | ".join(span_ranges) if span_ranges else ""
            if hdr:
                out_parts.append(f"\n% === MaViLS === {label} | {hdr}\n")
            else:
                out_parts.append(f"\n% === MaViLS === {label}\n")

            out_parts.append("\\paragraph*{ASR}\n\\begin{quote}\\small\n")

            # write ALL segments (no truncation)
            for (src_i, s) in segs_merged:
                t0 = fmt_time(s["t_start"])
                t1 = fmt_time(s["t_end"])
                idx = s.get("idx")
                txt = latex_escape(s.get("text", "") or "")
                prefix = f"[src={src_i} {t0}-{t1}]"
                if idx is not None:
                    prefix = f"[src={src_i} #{idx} {t0}-{t1}]"
                # each segment as its own paragraph line
                out_parts.append(f"\\noindent {latex_escape(prefix)} {txt}\\par\n")

            out_parts.append("\\end{quote}\n")

        last = insert_pos

    out_parts.append(tex[last:])
    return "".join(out_parts)

# -------------------- Main --------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in_tex", required=True)
    ap.add_argument("--xlsx", nargs="+", required=True)
    ap.add_argument("--srt", nargs="+", required=True)
    ap.add_argument("--out_tex", required=True)
    args = ap.parse_args()

    if len(args.xlsx) != len(args.srt):
        raise SystemExit(f"[ERR] xlsx count != srt count ({len(args.xlsx)} vs {len(args.srt)}). They must match.")

    tex, all_labels = read_tex_and_page_labels(args.in_tex)
    if not all_labels:
        raise SystemExit("[ERR] No page markers found in in_tex. Expected lines like: % <base>_page_<n>")

    # load SRTs
    srts: List[List[Dict]] = []
    video_ends: List[float] = []
    for srt_path in args.srt:
        segs = parse_srt(srt_path)
        srts.append(segs)
        ve = max((s["t_end"] for s in segs), default=0.0)
        video_ends.append(float(ve))

    # load XLSX and convert to spans
    spans_all: List[List[Dict]] = []
    max_values: List[int] = []
    steps: List[float] = []
    for i, xlsx_path in enumerate(args.xlsx):
        rows = read_kv_xlsx(xlsx_path)
        if not rows:
            print(f"[WARN] Empty xlsx: {xlsx_path}")
            spans_all.append([])
            max_values.append(0)
            steps.append(1.0)
            continue
        mv = max(v for _, v in rows)
        max_values.append(int(mv))
        spans, step = kv_to_spans(rows, video_end=video_ends[i] if i < len(video_ends) else None)
        spans_all.append(spans)
        steps.append(step)

    # allocate page label lists for each src
    page_lists = allocate_pages_for_sources(all_labels, max_values)

    # build mapping value->label and also label->span_ranges
    page2segs_allsrc: List[Dict[str, List[Dict]]] = []
    page2spaninfo_allsrc: List[Dict[str, List[Tuple[float, float]]]] = []

    total_written = 0
    total_off = 0

    for src_i in range(len(args.xlsx)):
        pages = page_lists[src_i] if src_i < len(page_lists) else []
        value_to_label = {i+1: pages[i] for i in range(len(pages))}

        # build label -> spans (for header info)
        label2spans: Dict[str, List[Tuple[float, float]]] = {}
        for sp in spans_all[src_i]:
            lab = value_to_label.get(int(sp["value"]))
            if not lab:
                continue
            label2spans.setdefault(lab, []).append((float(sp["t_start"]), float(sp["t_end"])))
        # assign segments
        page2segs, offslide = assign_segments_to_pages(srts[src_i], spans_all[src_i], value_to_label)

        written_here = sum(len(v) for v in page2segs.values())
        off_here = len(offslide)
        total_written += written_here
        total_off += off_here

        print(f"[SRC {src_i}] xlsx={args.xlsx[src_i]}")
        print(f"         spans={len(spans_all[src_i])}  step~{steps[src_i]:.3f}s  max_value={max_values[src_i]}")
        print(f"         pages_available={len(pages)}  written_to_pages={written_here}  offslide={off_here}")

        page2segs_allsrc.append(page2segs)
        page2spaninfo_allsrc.append(label2spans)

    stitched = stitch_into_tex(tex, page2segs_allsrc, page2spaninfo_allsrc)

    out_path = Path(args.out_tex)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(stitched, encoding="utf-8")

    print(f"[OK] wrote {args.out_tex}")
    print(f"[STAT] total_written_to_pages={total_written}  total_offslide={total_off}")

if __name__ == "__main__":
    main()
