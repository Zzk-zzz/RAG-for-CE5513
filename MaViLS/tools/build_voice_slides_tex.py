#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Voice-first builder:
- Keep ALL voice lines (never drop).
- For each voice cue, map (by timestamp) to a slide page index from MaViLS xlsx.
- Generate:
  - voice_text_body.tex (voice timeline w/ slide hyperlinks)
  - slides_text_body.tex (slides pages w/ hypertarget anchors)
  - display_merged.tex (single compileable tex: voice first, slides later)

Patch:
- Merge adjacent voice cues when:
    same lecture_label AND same slide_anchor (both None counts as same) AND gap <= --gap (default 1.0s)
- Allow running with NO CLI args by using hard-coded defaults below.
"""

from __future__ import annotations

import argparse
import bisect
import os
import re
import sys
from dataclasses import dataclass
from typing import List, Optional, Tuple

# ============================
# Defaults: EDIT THESE ONCE
# ============================
DEFAULT_SLIDES_TEX = r"E:\Code\outputs\voice_first\slides_text_body.tex"
DEFAULT_XLSX_LIST = [
    r"E:\Code\outputs\mavils_pair12_mean_matching_all_0comma2.xlsx",
    r"E:\Code\outputs\mavils_pair34_mean_matching_all_0comma2.xlsx",
]
DEFAULT_SRT_LIST = [
    r"E:\Code\outputs\lecture01.srt",
    r"E:\Code\outputs\lecture02.srt",
]
DEFAULT_LECTURE_LABELS = [
    "Lecture 01+02",
    "Lecture 03+04",
]
DEFAULT_OUT_DIR = r"E:\Code\outputs\voice_first"

DEFAULT_DEBOUNCE_S = 2.0   # suppress flicker in MaViLS sampling -> spans
DEFAULT_GAP_S = 1.0        # NEW: merge gap threshold for voice cues


# --- Optional deps
try:
    import pysrt  # type: ignore
except Exception:
    pysrt = None  # fallback to manual parser

try:
    import openpyxl  # type: ignore
except Exception as e:
    raise RuntimeError("Missing dependency: openpyxl. Install with `pip install openpyxl`") from e


# ----------------------------
# Data structures
# ----------------------------
@dataclass
class SlidePage:
    page_id: str          # raw marker string from merged slides tex
    anchor: str           # sanitized anchor for \hypertarget
    content: str          # tex content of this page (body slice)
    lecture_num: Optional[int]  # parsed lecture number if available


@dataclass
class Span:
    start: float          # seconds
    end: float            # seconds
    slide_idx: int        # 1-based index into the slide subset for this source


@dataclass
class VoiceCue:
    lecture_label: str
    start: float
    end: float
    text: str
    slide_anchor: Optional[str]
    slide_page_id: Optional[str]


# ----------------------------
# Utilities
# ----------------------------
_LATEX_SPECIALS = {
    "\\": r"\textbackslash{}",
    "&": r"\&",
    "%": r"\%",
    "$": r"\$",
    "#": r"\#",
    "_": r"\_",
    "{": r"\{",
    "}": r"\}",
    "~": r"\textasciitilde{}",
    "^": r"\textasciicircum{}",
}

def escape_latex(s: str) -> str:
    # remove basic html-ish tags that sometimes appear in srt
    s = re.sub(r"<[^>]+>", "", s)
    out = []
    for ch in s:
        out.append(_LATEX_SPECIALS.get(ch, ch))
    return "".join(out)

def time_hhmmssmmm(t: float) -> str:
    # t in seconds
    if t < 0:
        t = 0.0
    ms = int(round(t * 1000))
    hh = ms // 3600000
    ms -= hh * 3600000
    mm = ms // 60000
    ms -= mm * 60000
    ss = ms // 1000
    ms -= ss * 1000
    return f"{hh:02d}:{mm:02d}:{ss:02d}.{ms:03d}"

def sanitize_anchor(page_id: str) -> str:
    # keep it stable + latex-safe
    a = page_id.strip()
    a = re.sub(r"\s+", "_", a)
    a = re.sub(r"[^A-Za-z0-9_\-]+", "-", a)
    return "slide-" + a

def parse_lecture_num(page_id: str) -> Optional[int]:
    m = re.search(r"Lecture[_\s]*0*([0-9]+)", page_id, flags=re.IGNORECASE)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


# ----------------------------
# SRT parsing
# ----------------------------
def _parse_srt_time_to_seconds(s: str) -> float:
    # "HH:MM:SS,mmm"
    m = re.match(r"^\s*(\d+):(\d+):(\d+)[,\.](\d+)\s*$", s)
    if not m:
        raise ValueError(f"Bad time format: {s}")
    hh, mm, ss, ms = map(int, m.groups())
    return hh * 3600 + mm * 60 + ss + ms / 1000.0

def read_srt(path: str, encoding_hint: str = "utf-8") -> List[Tuple[float, float, str]]:
    if pysrt is not None:
        # try common encodings
        for enc in (encoding_hint, "utf-8-sig", "cp932", "gb18030"):
            try:
                subs = pysrt.open(path, encoding=enc)
                out = []
                for it in subs:
                    start = it.start.ordinal / 1000.0
                    end = it.end.ordinal / 1000.0
                    text = (it.text or "").replace("\n", " ").strip()
                    out.append((start, end, text))
                return out
            except Exception:
                continue

    # fallback manual parser
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        raw = f.read()

    blocks = re.split(r"\r?\n\r?\n+", raw.strip())
    cues = []
    for b in blocks:
        lines = [ln.strip("\r") for ln in b.splitlines() if ln.strip("\r").strip() != ""]
        if len(lines) < 2:
            continue
        # sometimes first line is index
        if re.match(r"^\d+$", lines[0]):
            lines = lines[1:]
        if not lines:
            continue
        # time line
        m = re.match(r"^(.+?)\s*-->\s*(.+?)$", lines[0])
        if not m:
            continue
        start = _parse_srt_time_to_seconds(m.group(1).strip())
        end = _parse_srt_time_to_seconds(m.group(2).strip())
        text = " ".join(lines[1:]).strip()
        cues.append((start, end, text))
    return cues


# ----------------------------
# XLSX mapping parsing
# ----------------------------
def read_mavils_xlsx_key_value(path: str) -> Tuple[List[float], List[int]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sh = wb.active

    # Expect header: Key | Value
    times: List[float] = []
    vals: List[int] = []
    for r in range(2, sh.max_row + 1):
        k = sh.cell(row=r, column=1).value
        v = sh.cell(row=r, column=2).value
        if k is None or v is None:
            continue
        try:
            times.append(float(k))
            vals.append(int(v))
        except Exception:
            continue

    wb.close()

    if not times:
        raise RuntimeError(f"No usable rows in xlsx: {path}")

    # ensure sorted by time
    pairs = sorted(zip(times, vals), key=lambda x: x[0])
    times = [p[0] for p in pairs]
    vals = [p[1] for p in pairs]
    return times, vals

def compress_to_spans(times: List[float], vals: List[int], duration_s: float, debounce_s: float) -> List[Span]:
    """
    Convert sampled mapping (time -> slide_idx) into piecewise-constant spans.
    Debounce: drop very short runs by merging them into previous value.
    """
    if len(times) != len(vals):
        raise ValueError("times/vals length mismatch")

    # Build initial runs (value, start, end)
    runs: List[Tuple[int, float, float]] = []
    cur_v = vals[0]
    cur_start = times[0]
    for i in range(1, len(times)):
        t = times[i]
        v = vals[i]
        if v != cur_v:
            runs.append((cur_v, cur_start, t))
            cur_v = v
            cur_start = t
    runs.append((cur_v, cur_start, max(duration_s, times[-1])))

    # Debounce: if a run is too short, merge into previous (simple + effective for flicker)
    if debounce_s > 0 and runs:
        cleaned: List[Tuple[int, float, float]] = []
        for v, s, e in runs:
            if cleaned and (e - s) < debounce_s:
                # merge short run into previous
                pv, ps, pe = cleaned[-1]
                cleaned[-1] = (pv, ps, e)
            else:
                cleaned.append((v, s, e))
        runs = cleaned

    # Merge any adjacent equal values (after debounce)
    merged: List[Tuple[int, float, float]] = []
    for v, s, e in runs:
        if not merged:
            merged.append((v, s, e))
            continue
        pv, ps, pe = merged[-1]
        if v == pv:
            merged[-1] = (pv, ps, e)
        else:
            merged.append((v, s, e))

    spans = [Span(start=s, end=e, slide_idx=v) for (v, s, e) in merged]
    return spans

def find_slide_for_time(spans: List[Span], t: float) -> Optional[int]:
    if not spans:
        return None
    starts = [sp.start for sp in spans]
    i = bisect.bisect_right(starts, t) - 1
    if i < 0:
        return None
    sp = spans[i]
    if t < sp.end:
        return sp.slide_idx
    return None


# ----------------------------
# Slides tex parsing
# ----------------------------
def read_tex_preamble_and_body(tex_path: str) -> Tuple[str, str]:
    with open(tex_path, "r", encoding="utf-8", errors="replace") as f:
        tex = f.read()

    m_begin = re.search(r"\\begin\{document\}", tex)
    m_end = re.search(r"\\end\{document\}", tex)
    if not m_begin or not m_end or m_end.start() <= m_begin.end():
        # fallback: treat whole file as body
        return "", tex

    preamble = tex[:m_begin.start()]
    body = tex[m_begin.end():m_end.start()]
    return preamble, body

def split_slides_body_into_pages(body: str) -> List[SlidePage]:
    """
    Split by marker lines like:
      % Lecture 01_xxx_page_6
    """
    # Capture the marker line as page_id
    pattern = re.compile(r"^\s*%\s*(Lecture.*?_page_\d+)\s*$", flags=re.IGNORECASE | re.MULTILINE)
    matches = list(pattern.finditer(body))

    if not matches:
        # no markers => single page
        pid = "SlidesBody_page_1"
        return [SlidePage(page_id=pid, anchor=sanitize_anchor(pid), content=body, lecture_num=parse_lecture_num(pid))]

    pages: List[SlidePage] = []
    for i, m in enumerate(matches):
        page_id = m.group(1).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
        content = body[start:end]
        pages.append(
            SlidePage(
                page_id=page_id,
                anchor=sanitize_anchor(page_id),
                content=content,
                lecture_num=parse_lecture_num(page_id),
            )
        )
    return pages

def ensure_hyperref_in_preamble(preamble: str) -> str:
    if "\\usepackage{hyperref}" in preamble:
        return preamble
    # insert near end of preamble
    return preamble.rstrip() + "\n\\usepackage{hyperref}\n"


# ----------------------------
# NEW: Merge voice cues
# ----------------------------
def merge_voice_cues(cues: List[VoiceCue], gap_s: float) -> List[VoiceCue]:
    """
    Merge adjacent cues when:
      - same lecture_label
      - same slide_anchor (both None counts as same)
      - next.start - prev.end <= gap_s
    Text is concatenated with a space.
    """
    if not cues:
        return []

    merged: List[VoiceCue] = [cues[0]]
    for c in cues[1:]:
        p = merged[-1]

        same_lecture = (c.lecture_label == p.lecture_label)
        same_anchor = (c.slide_anchor == p.slide_anchor)  # includes None==None
        close_enough = (c.start - p.end) <= gap_s

        if same_lecture and same_anchor and close_enough:
            # merge into p
            new_text = (p.text or "").strip()
            add_text = (c.text or "").strip()
            if new_text and add_text:
                new_text = new_text + " " + add_text
            else:
                new_text = new_text + add_text

            merged[-1] = VoiceCue(
                lecture_label=p.lecture_label,
                start=p.start,
                end=max(p.end, c.end),
                text=new_text,
                slide_anchor=p.slide_anchor,
                slide_page_id=p.slide_page_id,  # same_anchor implies same pageid in your inherited logic
            )
        else:
            merged.append(c)

    return merged


# ----------------------------
# LaTeX writers
# ----------------------------
def write_slides_text_body(out_path: str, pages: List[SlidePage]) -> None:
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("% Auto-generated slides text body (anchors per page)\n\n")
        for pg in pages:
            f.write(f"% {pg.page_id}\n")
            f.write(f"\\hypertarget{{{pg.anchor}}}{{}}\n")
            # keep page content as-is; just add a small header for debug navigation
            f.write(f"\\noindent\\textbf{{{escape_latex(pg.page_id)}}}\\\\\n")
            f.write(pg.content.strip() + "\n\n")
            f.write("\\clearpage\n\n")

def write_voice_text_body(out_path: str, cues: List[VoiceCue]) -> None:
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("% Auto-generated voice text body (never drops any cue)\n\n")

        current_lecture = None
        for c in cues:
            if c.lecture_label != current_lecture:
                current_lecture = c.lecture_label
                f.write(f"\\section*{{{escape_latex(current_lecture)}}}\n\n")

            tstr = f"{time_hhmmssmmm(c.start)}--{time_hhmmssmmm(c.end)}"
            if c.slide_anchor and c.slide_page_id:
                link = f"\\hyperlink{{{c.slide_anchor}}}{{\\texttt{{[Slide]}}}}"
                sid = escape_latex(c.slide_page_id)
                f.write(f"\\noindent\\texttt{{{tstr}}} {link} \\texttt{{{sid}}}\\\\\n")
            else:
                f.write(f"\\noindent\\texttt{{{tstr}}} \\texttt{{[OFFSLIDE]}}\\\\\n")

            txt = escape_latex(c.text.strip())
            if not txt:
                txt = "(empty)"
            f.write(txt + "\n\n")

def write_display_merged_tex(out_path: str, preamble: str, voice_body_filename: str, slides_body_filename: str) -> None:
    preamble = ensure_hyperref_in_preamble(preamble)

    # If preamble is empty (fallback), create a minimal one.
    if not preamble.strip():
        preamble = (
            "\\documentclass[11pt]{article}\n"
            "\\usepackage[margin=1in]{geometry}\n"
            "\\usepackage{hyperref}\n"
            "\\usepackage{graphicx}\n"
            "\\usepackage{longtable}\n"
        )

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(preamble.rstrip() + "\n")
        f.write("\\begin{document}\n\n")
        f.write("\\tableofcontents\n\\clearpage\n\n")

        # Voice-first display
        f.write("\\section{Voice transcript (voice-first)}\n")
        f.write(f"\\input{{{voice_body_filename}}}\n")
        f.write("\\clearpage\n\n")

        # Slides text after (so voice links jump forward)
        f.write("\\section{Slides text (anchors)}\n")
        f.write(f"\\input{{{slides_body_filename}}}\n")

        f.write("\n\\end{document}\n")


# ----------------------------
# Main pipeline
# ----------------------------
def build(
    slides_tex: str,
    xlsx_list: List[str],
    srt_list: List[str],
    lecture_labels: List[str],
    out_dir: str,
    debounce_s: float,
    gap_s: float,
) -> None:
    if len(xlsx_list) != len(srt_list) or len(srt_list) != len(lecture_labels):
        raise ValueError("Lengths must match: --xlsx, --srt, --lecture_label")

    os.makedirs(out_dir, exist_ok=True)

    preamble, body = read_tex_preamble_and_body(slides_tex)
    pages = split_slides_body_into_pages(body)

    # Determine where Lecture 03 starts (for splitting slide subsets)
    start_lecture3 = None
    for i, pg in enumerate(pages):
        if (pg.lecture_num is not None) and pg.lecture_num >= 3:
            start_lecture3 = i
            break

    # Read slide counts from xlsx (max value)
    max_vals = []
    for xlsx in xlsx_list:
        _, vals = read_mavils_xlsx_key_value(xlsx)
        max_vals.append(max(vals))

    # Build slide subsets per source
    subsets: List[List[SlidePage]] = []
    if len(xlsx_list) == 2:
        # default assumption: src0 -> Lecture 01&02; src1 -> Lecture 03&04
        split_i = start_lecture3 if start_lecture3 is not None else max_vals[0]
        subsets.append(pages[:split_i])
        subsets.append(pages[split_i: split_i + max_vals[1]])
    else:
        # generic: sequential slicing by each src max slide index
        cursor = 0
        for mv in max_vals:
            subsets.append(pages[cursor: cursor + mv])
            cursor += mv

    # Generate slides body (all pages, anchors global)
    slides_body_path = os.path.join(out_dir, "slides_text_body.tex")
    write_slides_text_body(slides_body_path, pages)

    all_voice_cues: List[VoiceCue] = []

    # Process each lecture recording
    for src_idx, (xlsx_path, srt_path, lec_label) in enumerate(zip(xlsx_list, srt_list, lecture_labels)):
        srt_cues = read_srt(srt_path)
        duration = 0.0
        for st, en, _ in srt_cues:
            duration = max(duration, en)

        times, vals = read_mavils_xlsx_key_value(xlsx_path)
        spans = compress_to_spans(times, vals, duration_s=duration, debounce_s=debounce_s)

        slide_subset = subsets[src_idx]
        last_known_anchor: Optional[str] = None
        last_known_pageid: Optional[str] = None

        for st, en, txt in srt_cues:
            t_pick = 0.5 * (st + en)  # use mid-time
            slide_idx_1based = find_slide_for_time(spans, t_pick)

            anchor = None
            pageid = None
            if slide_idx_1based is not None and 1 <= slide_idx_1based <= len(slide_subset):
                pg = slide_subset[slide_idx_1based - 1]
                anchor = pg.anchor
                pageid = pg.page_id
                last_known_anchor = anchor
                last_known_pageid = pageid
            else:
                # Keep voice, never drop.
                # If not matched, inherit previous known slide within this lecture (as you agreed).
                anchor = last_known_anchor
                pageid = last_known_pageid

            all_voice_cues.append(
                VoiceCue(
                    lecture_label=lec_label,
                    start=st,
                    end=en,
                    text=txt,
                    slide_anchor=anchor,
                    slide_page_id=pageid,
                )
            )

        # quick stats per lecture
        matched = sum(1 for c in all_voice_cues if c.lecture_label == lec_label and c.slide_anchor is not None)
        total = sum(1 for c in all_voice_cues if c.lecture_label == lec_label)
        print(f"[{lec_label}] cues={total}, matched_or_inherited={matched}, offslide={total - matched}")

    # NEW: merge adjacent voice cues
    all_voice_cues = merge_voice_cues(all_voice_cues, gap_s=gap_s)

    voice_body_path = os.path.join(out_dir, "voice_text_body.tex")
    write_voice_text_body(voice_body_path, all_voice_cues)

    display_path = os.path.join(out_dir, "display_merged.tex")
    write_display_merged_tex(
        display_path,
        preamble=preamble,
        voice_body_filename="voice_text_body.tex",
        slides_body_filename="slides_text_body.tex",
    )

    print("\nDone.")
    print("Generated:")
    print(f"  - {voice_body_path}")
    print(f"  - {slides_body_path}")
    print(f"  - {display_path}")


def main():
    # If no CLI args: use hard-coded defaults (you requested this).
    if len(sys.argv) == 1:
        build(
            slides_tex=DEFAULT_SLIDES_TEX,
            xlsx_list=DEFAULT_XLSX_LIST,
            srt_list=DEFAULT_SRT_LIST,
            lecture_labels=DEFAULT_LECTURE_LABELS,
            out_dir=DEFAULT_OUT_DIR,
            debounce_s=DEFAULT_DEBOUNCE_S,
            gap_s=DEFAULT_GAP_S,
        )
        return

    ap = argparse.ArgumentParser()
    ap.add_argument("--slides_tex", required=True, help="Path to merged slides tex OR slides_text_body.tex with % Lecture..._page_ markers")
    ap.add_argument("--xlsx", nargs="+", required=True, help="List of MaViLS xlsx mappings (same count as --srt)")
    ap.add_argument("--srt", nargs="+", required=True, help="List of SRT files (same count as --xlsx)")
    ap.add_argument("--lecture_label", nargs="+", required=True, help="Labels for each SRT (e.g., Lecture 01+02, Lecture 03+04)")
    ap.add_argument("--out_dir", required=True, help="Output directory")
    ap.add_argument("--debounce", type=float, default=2.0, help="Debounce seconds to suppress flicker (default 2.0)")
    ap.add_argument("--gap", type=float, default=1.0, help="Merge voice cues gap seconds (default 1.0)")
    args = ap.parse_args()

    build(
        slides_tex=args.slides_tex,
        xlsx_list=args.xlsx,
        srt_list=args.srt,
        lecture_labels=args.lecture_label,
        out_dir=args.out_dir,
        debounce_s=args.debounce,
        gap_s=args.gap,
    )

if __name__ == "__main__":
    main()
